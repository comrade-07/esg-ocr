from datetime import date, datetime

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.pipeline.run_silver_pipeline import (
    SILVER_AGGREGATED_COLUMNS,
    SILVER_BUSINESS_MAPPING_FIXED_COLUMNS,
    SILVER_CURATED_COLUMNS,
    SILVER_PRORATION_CALCULATION_COLUMNS,
    SILVER_PRORATION_SPLIT_COLUMNS,
    SILVER_PRORATED_COLUMNS,
    SILVER_TEMPLATE_PREPARATION_COLUMNS,
    MANUAL_DATA_ENTRY_HELPER_COLUMN,
    MANUAL_DATA_ENTRY_HELPER_VALUE,
    _with_data_quality_column,
    build_silver_aggregated,
    build_silver_business_mapping,
    add_consumption_quantity_unit,
    build_manual_data_entry_proration_input,
    build_silver_curated,
    build_silver_proration_calculation,
    build_silver_proration_split,
    build_silver_prorated,
    build_silver_template_preparation,
    build_silver_template_output,
    legacy_silver_output_filename,
    run_silver_pipeline,
    silver_output_filename,
)


def _as_date(value):
    if isinstance(value, datetime):
        return value.date()
    return value


def _without_numbered_sharepoint_columns(columns):
    return [
        column
        for column in columns
        if not (
            isinstance(column, str)
            and column.startswith("sharepoint_link_")
            and column.rsplit("_", 1)[1].isdigit()
        )
    ]


def test_data_quality_column_is_inserted_after_document_confidence():
    source_df = pd.DataFrame([
        {"source_file": "ocr.json", "document_confidence": "0.99", "status": "succeeded"},
        {"source_file": "manual.csv", "document_confidence": "", "data_quality": "Manual"},
    ])

    result = _with_data_quality_column(source_df)

    assert list(result.columns) == [
        "source_file",
        "document_confidence",
        "data_quality",
        "status",
    ]
    assert result.loc[0, "data_quality"] == "Actual"
    assert result.loc[1, "data_quality"] == "Manual"


def test_silver_pipeline_transforms_approved_review_rows(tmp_path):
    config_dir = tmp_path / "config"
    bronze_dir = tmp_path / "bronze"
    silver_dir = tmp_path / "silver"
    checkpoint_dir = tmp_path / "checkpoints"
    mappings_dir = tmp_path / "mappings"
    config_dir.mkdir()
    bronze_dir.mkdir()
    checkpoint_dir.mkdir()
    mappings_dir.mkdir()
    (config_dir / "settings.yaml").write_text(
        "\n".join([
            "paths:",
            f"  bronze_output: {bronze_dir.as_posix()}",
            f"  silver_excel_output: {silver_dir.as_posix()}",
            f"  review_checkpoint_output: {checkpoint_dir.as_posix()}",
            "",
        ]),
        encoding="utf-8",
    )
    (config_dir / "mapping_files.yaml").write_text(
        "\n".join([
            "mapping_files:",
            "  mapping:",
            f"    path: {mappings_dir.joinpath('mapping.xlsx').as_posix()}",
            "    sheet_name: mapping",
            "    ocr_key_fields: [account_number]",
            "    workbook_key_columns: [account_number]",
            "  energy_source_allocation:",
            f"    path: {mappings_dir.joinpath('energy_source_allocation.xlsx').as_posix()}",
            "    sheet_name: energy_source_allocation",
            "    delimiter: _",
            "    ocr_key_fields: [unit_name, supplier, start_date]",
            "    workbook_key_columns: [unit_name, supplier_name, start_date]",
            "    date_key_fields: [start_date]",
            "  contracts:",
            f"    path: {mappings_dir.joinpath('contracts.xlsx').as_posix()}",
            "    sheet_name: contracts",
            "    delimiter: _",
            "    ocr_key_fields: [unit_name, supplier, contract_start_date]",
            "    workbook_key_columns: [unit_name, supplier_name, contract_start_date]",
            "    date_key_fields: [contract_start_date]",
            "",
        ]),
        encoding="utf-8",
    )
    pd.DataFrame(
        [
            {
                "active": "Yes",
                "account_number": "ACC-001",
                "division": "Tanker",
                "legal_entity": "Mapped_Legal Entity One",
                "unit_name": "Mapped_Unit One",
                "supplier_name": "Mapped_Supplier One",
                "ocr_unit_name_list": "",
                "ocr_legal_entity_name_list": "",
                "division_shorthand": "TNK",
                "facility_type": "Terminal",
                "scope": "Scope 2",
                "facility_identifier": "FAC-001",
                "activity_group": "Electricity",
                "invoice_count": 1,
                "invoice_frequency": "Monthly",
                "consumption_unit": "kWh",
                "decimal_separator": ",",
                "currency": "USD",
            },
            {
                "active": "Yes",
                "account_number": "ACC-002",
                "division": "Seafarm",
                "legal_entity": "Mapped Legal Entity Two",
                "unit_name": "Mapped Unit Two",
                "supplier_name": "Mapped Supplier Two",
                "ocr_unit_name_list": "Bogota Alias;Unit Alias Two",
                "ocr_legal_entity_name_list": "",
                "division_shorthand": "SF",
                "facility_type": "Office",
                "scope": "Scope 2",
                "facility_identifier": "FAC-002",
                "activity_group": "Electricity",
                "invoice_count": 2,
                "invoice_frequency": "Quarterly",
                "consumption_unit": "MWh",
                "decimal_separator": ".",
                "currency": "EUR",
            },
            {
                "active": "Yes",
                "account_number": "ACC-003",
                "division": "Tankers",
                "legal_entity": "Mapped Legal Entity Three",
                "unit_name": "Mapped Unit Three",
                "supplier_name": "Mapped Supplier Three",
                "ocr_unit_name_list": "",
                "ocr_legal_entity_name_list": "Legal Alias Three|Another Legal Alias",
                "division_shorthand": "TKS",
                "facility_type": "Warehouse",
                "scope": "Scope 1",
                "facility_identifier": "FAC-003",
                "activity_group": "Fuel",
                "invoice_count": 3,
                "invoice_frequency": "Annual",
                "consumption_unit": "kWh",
                "decimal_separator": ",",
                "currency": "BRL",
            },
            {
                "active": "Yes",
                "account_number": "ACC-MANUAL",
                "division": "Tanker",
                "legal_entity": "Manual Legal",
                "unit_name": "Manual Unit",
                "supplier_name": "Manual Supplier",
                "ocr_unit_name_list": "Manual Unit",
                "ocr_legal_entity_name_list": "Manual Legal",
                "division_shorthand": "TNK",
                "facility_type": "Terminal",
                "scope": "Scope 2",
                "facility_identifier": "FAC-MANUAL",
                "activity_group": "Electricity",
                "invoice_count": 1,
                "invoice_frequency": "Monthly",
                "consumption_unit": "kWh",
                "decimal_separator": ".",
                "currency": "USD",
            },
        ]
    ).to_excel(mappings_dir / "mapping.xlsx", sheet_name="mapping", index=False)
    pd.DataFrame([
        {
            "lookup_key": "mapped_unit one_mapped_supplier one_2025-06-01",
            "active": "Yes",
            "unit_name": "Mapped_Unit One",
            "supplier_name": "Mapped_Supplier One",
            "start_date": "2025-06-01",
            "fossil_fuel_%": 0.294,
            "renewable_energy_%": 0.507,
            "nuclear_%": 0.199,
        },
        {
            "lookup_key": "manual unit_manual supplier_2025-06-01",
            "active": "Yes",
            "unit_name": "Manual Unit",
            "supplier_name": "Manual Supplier",
            "start_date": "2025-06-01",
            "fossil_fuel_%": 0.2,
            "renewable_energy_%": 0.7,
            "nuclear_%": 0.1,
        },
        {
            "lookup_key": "manual unit_manual supplier_2025-07-01",
            "active": "Yes",
            "unit_name": "Manual Unit",
            "supplier_name": "Manual Supplier",
            "start_date": "2025-07-01",
            "fossil_fuel_%": 0.25,
            "renewable_energy_%": 0.65,
            "nuclear_%": 0.1,
        },
    ]).to_excel(
        mappings_dir / "energy_source_allocation.xlsx",
        sheet_name="energy_source_allocation",
        index=False,
    )
    pd.DataFrame([
        {
            "lookup_key": "mapped_unit one_mapped_supplier one_2025-06-01",
            "active": "Yes",
            "unit_name": "Mapped_Unit One",
            "supplier_name": "Mapped_Supplier One",
            "contract_start_date": "2025-06-01",
            "contractual_instruments": "Contracts(e.g. PPAs)",
        },
        {
            "lookup_key": "manual unit_manual supplier_2025-06-01",
            "active": "Yes",
            "unit_name": "Manual Unit",
            "supplier_name": "Manual Supplier",
            "contract_start_date": "2025-06-01",
            "contractual_instruments": "Manual green tariff",
        },
        {
            "lookup_key": "manual unit_manual supplier_2025-07-01",
            "active": "Yes",
            "unit_name": "Manual Unit",
            "supplier_name": "Manual Supplier",
            "contract_start_date": "2025-07-01",
            "contractual_instruments": "Manual green tariff",
        },
    ]).to_excel(
        mappings_dir / "contracts.xlsx",
        sheet_name="contracts",
        index=False,
    )
    approved_file = checkpoint_dir / "step_5_approved_silver_checkpoint.csv"
    legacy_normalized_output = silver_dir / legacy_silver_output_filename("scope2", "normalized")
    stale_numbered_prorated_output = silver_dir / "06_scope2_silver_prorated.xlsx"
    legacy_normalized_output.parent.mkdir(parents=True)
    legacy_normalized_output.write_text("old duplicate output", encoding="utf-8")
    stale_numbered_prorated_output.write_text("old numbered output", encoding="utf-8")
    sharepoint_link = "https://example.sharepoint.com/sites/invoices/sample.pdf"
    approved_file.write_text(
        "\n".join([
            "source_file,account_number,legal_entity,unit_name,supplier,account_number_confidence,invoice_date,total_amount,quantity_1,quantity_2,quantity_3,quantity_4,quantity_5,quantity_6,consumption_start_date_1,consumption_end_date_1,consumption_start_date_2,consumption_end_date_2,createdDateTime,sharepoint_link,approval_status,manual_review_tag,manual_review_decision_count,total_amount_review_tag,total_amount_review_decision",
            f'one.json,ACC-001, STOLT-NIELSEN MIDDLE EAST DMCC ,"(Durban)"," Dubai Electricity & Water Authority",0.95,25 Jul 2025,"USD 1.234,56","1.234,56","123,45","1.234",,,0,1/6/2025,6/30/2025,,,2026-06-14T15:07:06Z,{sharepoint_link},AUTO_APPROVED,,0,,',
            'two.json,,STOLT_TANK CONTAINERS COLOMBIA SAS,Unit Alias Two,"Enel Colombia S.A. E.S.P.",,114/07/25,"€1,234.56","1,234.56",1234.56,,,,,01.07.25,01.08.25,,,2026-06-15T15:07:06Z,,MANUALLY_APPROVED,MANUALLY_REVIEWED,1,MANUALLY_REVIEWED,CORRECTED',
            'three.json,,Legal Alias Three,Unknown Unit,Original Supplier,,2025-08-01,"BRL 2.500,5","2.500,5",,,,,,2025-08-01,2025-08-31,,,2026-06-16T15:07:06Z,,AUTO_APPROVED,,0,,',
            'four.json,,Original Legal Entity,Original Unit,Original Supplier,,2025-09-01,"USD 9,876.5","9,876.5",,,,,,2025-09-01,2025-09-30,,,2026-06-17T15:07:06Z,,AUTO_APPROVED,,0,,',
            "",
        ]),
        encoding="utf-8",
    )
    (checkpoint_dir / "step_0_manual_data_entry_decisions_checkpoint.csv").write_text(
        "\n".join([
            "invoice_id,line_id,source_file,sharepoint_link,manual_entry_status,reviewed_by,reviewed_at,review_comment,data_quality,division,legal_entity_name,unit,consumption_start_date,consumption_end_date,transaction_date,amount_of_energy_consumed,energy_unit",
            "manual-one,1,manual.pdf,,COMPLETED,Reviewer,2026-06-30T00:00:00+00:00,,Estimated,Tanker,Manual Legal,Manual Unit,2025-06-15,2025-07-14,2025-07-20,300,kWh",
            "",
        ]),
        encoding="utf-8",
    )

    output = run_silver_pipeline(config_dir=config_dir)

    assert not legacy_normalized_output.exists()
    assert not stale_numbered_prorated_output.exists()

    reviewed_output = silver_dir / silver_output_filename("scope2", "reviewed")
    reviewed_workbook = load_workbook(reviewed_output)
    reviewed_worksheet = reviewed_workbook["SilverReviewed"]
    reviewed_headers = [cell.value for cell in reviewed_worksheet[1]]
    reviewed_rows = [
        dict(zip(reviewed_headers, row))
        for row in reviewed_worksheet.iter_rows(min_row=2, values_only=True)
    ]

    assert reviewed_output.exists()
    assert reviewed_rows[0]["account_number"] == "ACC-001"
    assert reviewed_rows[0]["legal_entity"] == " STOLT-NIELSEN MIDDLE EAST DMCC "
    assert reviewed_rows[0]["unit_name"] == "(Durban)"
    assert reviewed_rows[0]["supplier"] == " Dubai Electricity & Water Authority"
    assert reviewed_rows[0]["invoice_date"] == "25 Jul 2025"
    assert reviewed_rows[0]["data_quality"] == "Actual"
    assert "invoice_date_normalized" not in reviewed_rows[0]

    workbook = load_workbook(output)
    worksheet = workbook["SilverNormalized"]
    headers = [cell.value for cell in worksheet[1]]
    rows = [
        dict(zip(headers, row))
        for row in worksheet.iter_rows(min_row=2, values_only=True)
    ]

    assert output.suffix == ".xlsx"
    assert len(rows) == 4
    assert rows[0]["source_file"] == "one.json"
    assert rows[0]["data_quality"] == "Actual"
    assert rows[0]["approval_status"] == "AUTO_APPROVED"
    assert rows[0]["sharepoint_link"] == sharepoint_link
    assert rows[0]["legal_entity"] == "Mapped_Legal Entity One"
    assert rows[0]["unit_name"] == "Mapped_Unit One"
    assert rows[0]["supplier_name"] == "Mapped_Supplier One"
    assert rows[0]["division"] == "Tanker"
    assert rows[0]["division_shorthand"] == "TNK"
    assert rows[0]["facility_type"] == "Terminal"
    assert rows[0]["scope"] == "Scope 2"
    assert rows[0]["facility_identifier"] == "FAC-001"
    assert rows[0]["activity_group"] == "Electricity"
    assert rows[0]["invoice_count"] == 1
    assert rows[0]["invoice_frequency"] == "Monthly"
    assert rows[0]["consumption_unit"] == "kWh"
    assert rows[0]["consumption_quantity_unit"] == "KWH"
    assert rows[0]["decimal_separator"] == ","
    assert rows[0]["currency"] == "USD"
    assert rows[0]["total_amount"] == 1234.56
    assert rows[0]["quantity_1"] == 1234.56
    assert rows[0]["quantity_2"] == 123.45
    assert rows[0]["quantity_3"] == 1234
    assert rows[0]["quantity_6"] == 0
    assert rows[0]["supplier"] == "DUBAI ELECTRICITY & WATER AUTHORITY"
    assert rows[0]["invoice_date"] == "25 Jul 2025"
    assert _as_date(rows[0]["invoice_date_normalized"]) == date(2025, 7, 25)
    assert _as_date(rows[0]["consumption_start_date_1_normalized"]) == date(2025, 6, 1)
    assert _as_date(rows[0]["consumption_end_date_1_normalized"]) == date(2025, 6, 30)
    assert rows[1]["source_file"] == "two.json"
    assert rows[1]["approval_status"] == "MANUALLY_APPROVED"
    assert rows[1]["legal_entity"] == "Mapped Legal Entity Two"
    assert rows[1]["unit_name"] == "Mapped Unit Two"
    assert rows[1]["supplier_name"] == "Mapped Supplier Two"
    assert rows[1]["invoice_count"] == 2
    assert rows[1]["invoice_frequency"] == "Quarterly"
    assert rows[1]["consumption_unit"] == "MWh"
    assert rows[1]["consumption_quantity_unit"] == "MWH"
    assert rows[1]["decimal_separator"] == "."
    assert rows[1]["currency"] == "EUR"
    assert rows[1]["quantity_1"] == 1234.56
    assert rows[1]["quantity_2"] == 1234.56
    assert rows[1]["supplier"] == "ENEL COLOMBIA S.A. E.S.P"
    assert rows[1]["invoice_date"] == "114/07/25"
    assert rows[1]["total_amount"] == 1234.56
    assert rows[1]["total_amount_review_decision"] == "CORRECTED"
    assert _as_date(rows[1]["invoice_date_normalized"]) == date(2025, 7, 25)
    assert rows[2]["legal_entity"] == "Mapped Legal Entity Three"
    assert rows[2]["unit_name"] == "Mapped Unit Three"
    assert rows[2]["supplier_name"] == "Mapped Supplier Three"
    assert rows[2]["invoice_count"] == 3
    assert rows[2]["invoice_frequency"] == "Annual"
    assert rows[2]["consumption_unit"] == "kWh"
    assert rows[2]["consumption_quantity_unit"] == "KWH"
    assert rows[2]["decimal_separator"] == ","
    assert rows[2]["currency"] == "BRL"
    assert rows[2]["total_amount"] == 2500.5
    assert rows[2]["quantity_1"] == 2500.5
    assert rows[3]["legal_entity"] == "Original Legal Entity"
    assert rows[3]["unit_name"] == "Original Unit"
    assert rows[3]["supplier_name"] is None
    assert rows[3]["invoice_count"] is None
    assert rows[3]["invoice_frequency"] is None
    assert rows[3]["consumption_unit"] is None
    assert rows[3]["consumption_quantity_unit"] is None
    assert rows[3]["decimal_separator"] is None
    assert rows[3]["currency"] is None
    assert rows[3]["total_amount"] == 9876.5
    assert rows[3]["quantity_1"] == 9876.5
    assert "createdDateTime_normalized" not in rows[0]

    account_index = headers.index("account_number")
    assert headers[account_index + 1] == "account_number_confidence"
    assert rows[0]["account_number_confidence"] == "0.95"
    division_index = headers.index("division")
    assert headers[division_index:division_index + 15] == [
        "division",
        "legal_entity",
        "unit_name",
        "supplier_name",
        "division_shorthand",
        "facility_type",
        "scope",
        "facility_identifier",
        "activity_group",
        "invoice_count",
        "invoice_frequency",
        "consumption_unit",
        "consumption_quantity_unit",
        "decimal_separator",
        "currency",
    ]

    normalized_column_index = headers.index("invoice_date_normalized") + 1
    assert worksheet.cell(row=2, column=normalized_column_index).number_format == "yyyy-mm-dd"

    sharepoint_column_index = headers.index("sharepoint_link") + 1
    sharepoint_cell = worksheet.cell(row=2, column=sharepoint_column_index)
    blank_sharepoint_cell = worksheet.cell(row=3, column=sharepoint_column_index)
    assert sharepoint_cell.hyperlink.target == sharepoint_link
    assert blank_sharepoint_cell.hyperlink is None

    curated_output = silver_dir / silver_output_filename("scope2", "curated")
    aggregated_output = silver_dir / silver_output_filename("scope2", "aggregated")
    proration_calculation_output = silver_dir / silver_output_filename("scope2", "proration_calculation")
    proration_split_output = silver_dir / silver_output_filename("scope2", "proration_split")
    prorated_output = silver_dir / silver_output_filename("scope2", "prorated")
    manual_proration_calculation_output = silver_dir / silver_output_filename("scope2", "manual_proration_calculation")
    manual_proration_split_output = silver_dir / silver_output_filename("scope2", "manual_proration_split")
    manual_prorated_output = silver_dir / silver_output_filename("scope2", "manual_prorated")
    business_mapping_output = silver_dir / silver_output_filename("scope2", "business_mapping")
    template_preparation_output = silver_dir / silver_output_filename("scope2", "template_preparation")
    template_output_output = silver_dir / silver_output_filename("scope2", "template_output")
    curated_workbook = load_workbook(curated_output)
    curated_worksheet = curated_workbook["SilverCurated"]
    curated_headers = [cell.value for cell in curated_worksheet[1]]

    assert curated_output.exists()
    assert curated_headers == SILVER_CURATED_COLUMNS
    assert curated_headers[curated_headers.index("data_quality") + 1] == "division"
    assert curated_worksheet.cell(row=2, column=curated_headers.index("data_quality") + 1).value == "Actual"
    assert curated_headers[curated_headers.index("unit_name") + 1] == "account_number"
    assert curated_headers[curated_headers.index("invoice_count") + 1] == "source_file"
    assert curated_headers[curated_headers.index("source_file") + 1] == "sharepoint_link"
    assert curated_headers[curated_headers.index("quantity_unit_1") + 1] == "consumption_unit"
    assert curated_headers[curated_headers.index("consumption_unit") + 1] == "consumption_quantity_unit"
    assert curated_headers[curated_headers.index("quantity_unit_2") + 1] == "total_amount"
    total_amount_column_index = curated_headers.index("total_amount") + 1
    quantity_1_column_index = curated_headers.index("quantity_1") + 1
    assert curated_worksheet.cell(row=2, column=total_amount_column_index).value == 1234.56
    assert curated_worksheet.cell(row=2, column=total_amount_column_index).data_type == "n"
    assert curated_worksheet.cell(row=2, column=total_amount_column_index).number_format == "General"
    assert curated_worksheet.cell(row=2, column=quantity_1_column_index).value == 1234.56
    assert curated_worksheet.cell(row=2, column=quantity_1_column_index).data_type == "n"
    assert curated_worksheet.cell(row=2, column=quantity_1_column_index).number_format == "General"
    assert curated_worksheet.cell(row=2, column=curated_headers.index("legal_entity") + 1).value == "Mapped_Legal Entity One"
    assert curated_worksheet.cell(row=2, column=curated_headers.index("unit_name") + 1).value == "Mapped_Unit One"
    assert curated_worksheet.cell(row=2, column=curated_headers.index("account_number") + 1).value == "ACC-001"
    assert curated_worksheet.cell(row=2, column=curated_headers.index("supplier_name") + 1).value == "Mapped_Supplier One"
    assert _as_date(curated_worksheet.cell(row=2, column=curated_headers.index("consumption_start_date_1_normalized") + 1).value) == date(2025, 6, 1)
    assert _as_date(curated_worksheet.cell(row=2, column=curated_headers.index("consumption_end_date_1_normalized") + 1).value) == date(2025, 6, 30)
    assert curated_worksheet.cell(row=2, column=curated_headers.index("source_file") + 1).value == "one.json"
    assert curated_worksheet.cell(row=2, column=curated_headers.index("sharepoint_link") + 1).value == sharepoint_link

    aggregated_workbook = load_workbook(aggregated_output)
    aggregated_worksheet = aggregated_workbook["SilverAggregated"]
    aggregated_headers = [cell.value for cell in aggregated_worksheet[1]]
    aggregated_quantity_1_column_index = aggregated_headers.index("aggregated_quantity_1") + 1
    aggregated_quantity_2_column_index = aggregated_headers.index("aggregated_quantity_2") + 1

    assert aggregated_output.exists()
    assert aggregated_headers == SILVER_AGGREGATED_COLUMNS
    assert aggregated_headers[aggregated_headers.index("data_quality") + 1] == "division"
    assert aggregated_worksheet.cell(row=2, column=aggregated_headers.index("data_quality") + 1).value == "Actual"
    assert aggregated_worksheet.cell(row=2, column=aggregated_headers.index("account_number") + 1).value == "ACC-001"
    assert aggregated_headers[aggregated_headers.index("invoice_count") + 1] == "source_file"
    assert aggregated_headers[aggregated_headers.index("source_file") + 1] == "sharepoint_link"
    assert "quantity_1" not in aggregated_headers
    assert "quantity_6" not in aggregated_headers
    assert aggregated_worksheet.cell(row=2, column=aggregated_headers.index("source_file") + 1).value == "one.json"
    assert aggregated_worksheet.cell(row=2, column=aggregated_headers.index("sharepoint_link") + 1).value == sharepoint_link
    assert _as_date(aggregated_worksheet.cell(row=2, column=aggregated_headers.index("consumption_start_date_1_normalized") + 1).value) == date(2025, 6, 1)
    assert _as_date(aggregated_worksheet.cell(row=2, column=aggregated_headers.index("consumption_end_date_1_normalized") + 1).value) == date(2025, 6, 30)
    assert aggregated_worksheet.cell(row=2, column=aggregated_quantity_1_column_index).value == 2592.01
    assert aggregated_worksheet.cell(row=2, column=aggregated_quantity_1_column_index).data_type == "n"
    assert aggregated_worksheet.cell(row=2, column=aggregated_quantity_2_column_index).value is None

    proration_calculation_workbook = load_workbook(proration_calculation_output)
    proration_calculation_worksheet = proration_calculation_workbook["SilverProrationCalculation"]
    proration_calculation_headers = [cell.value for cell in proration_calculation_worksheet[1]]

    assert proration_calculation_output.exists()
    assert proration_calculation_headers == SILVER_PRORATION_CALCULATION_COLUMNS
    assert proration_calculation_headers[proration_calculation_headers.index("data_quality") + 1] == "division"
    assert proration_calculation_worksheet.cell(
        row=2,
        column=proration_calculation_headers.index("data_quality") + 1,
    ).value == "Actual"
    assert proration_calculation_worksheet.cell(
        row=2,
        column=proration_calculation_headers.index("account_number") + 1,
    ).value == "ACC-001"

    proration_split_workbook = load_workbook(proration_split_output)
    proration_split_worksheet = proration_split_workbook["SilverProrationSplit"]
    proration_split_headers = [cell.value for cell in proration_split_worksheet[1]]

    assert proration_split_output.exists()
    assert proration_split_headers == SILVER_PRORATION_SPLIT_COLUMNS
    assert proration_split_headers[proration_split_headers.index("data_quality") + 1] == "division"
    assert proration_split_worksheet.cell(
        row=2,
        column=proration_split_headers.index("data_quality") + 1,
    ).value == "Actual"
    assert proration_split_worksheet.cell(
        row=2,
        column=proration_split_headers.index("account_number") + 1,
    ).value == "ACC-001"
    proration_split_number_columns = [
        "aggregate_quantity1_to_6",
        "average_consumption",
        "date_split_days_proration",
        "date_total_days_proration",
        "quantity_proration",
    ]
    for column_name in proration_split_number_columns:
        cell = proration_split_worksheet.cell(
            row=2,
            column=proration_split_headers.index(column_name) + 1,
        )
        assert cell.data_type == "n"
        assert cell.number_format == "General"

    prorated_workbook = load_workbook(prorated_output)
    prorated_worksheet = prorated_workbook["SilverProrated"]
    prorated_headers = [cell.value for cell in prorated_worksheet[1]]

    assert prorated_output.exists()
    assert _without_numbered_sharepoint_columns(prorated_headers) == SILVER_PRORATED_COLUMNS
    assert "sharepoint_link_1" in prorated_headers
    assert prorated_headers[prorated_headers.index("data_quality") + 1] == "division"
    assert prorated_worksheet.cell(row=2, column=prorated_headers.index("data_quality") + 1).value == "Actual"
    sharepoint_link_1_cell = prorated_worksheet.cell(row=2, column=prorated_headers.index("sharepoint_link_1") + 1)
    assert sharepoint_link_1_cell.value == sharepoint_link
    assert sharepoint_link_1_cell.hyperlink.target == sharepoint_link

    assert business_mapping_output.exists()
    business_mapping_workbook = load_workbook(business_mapping_output)
    business_mapping_worksheet = business_mapping_workbook["SilverBusinessMapping"]
    business_mapping_headers = [cell.value for cell in business_mapping_worksheet[1]]

    assert business_mapping_headers == SILVER_BUSINESS_MAPPING_FIXED_COLUMNS + ["sharepoint_link_1", "sharepoint_link_2"]
    assert business_mapping_headers[business_mapping_headers.index("Energy Unit") + 1] == "lookup_key"
    assert business_mapping_headers[business_mapping_headers.index("lookup_key") + 1] == "fossil_fuel_%"
    assert business_mapping_headers[business_mapping_headers.index("nuclear_%") + 1] == "contractual_instruments"
    assert business_mapping_headers[business_mapping_headers.index("contractual_instruments") + 1] == "Account number"
    assert business_mapping_worksheet.cell(row=2, column=business_mapping_headers.index("Data Quality") + 1).value == "Actual"
    assert business_mapping_worksheet.cell(row=2, column=business_mapping_headers.index("KPI Component") + 1).value == "Purchased electricity"
    assert _as_date(business_mapping_worksheet.cell(row=2, column=business_mapping_headers.index("Transaction Date") + 1).value) == date(2026, 6, 14)
    assert business_mapping_worksheet.cell(row=2, column=business_mapping_headers.index("lookup_key") + 1).value == "mapped_unit one_mapped_supplier one_2025-06-01"
    assert business_mapping_worksheet.cell(row=2, column=business_mapping_headers.index("fossil_fuel_%") + 1).value == 0.294
    assert business_mapping_worksheet.cell(row=2, column=business_mapping_headers.index("renewable_energy_%") + 1).value == 0.507
    assert business_mapping_worksheet.cell(row=2, column=business_mapping_headers.index("nuclear_%") + 1).value == 0.199
    assert business_mapping_worksheet.cell(row=2, column=business_mapping_headers.index("contractual_instruments") + 1).value == "Contracts(e.g. PPAs)"
    business_mapping_sharepoint_cell = business_mapping_worksheet.cell(
        row=2,
        column=business_mapping_headers.index("sharepoint_link_1") + 1,
    )
    assert business_mapping_sharepoint_cell.value == sharepoint_link
    assert business_mapping_sharepoint_cell.hyperlink.target == sharepoint_link

    assert template_preparation_output.exists()
    template_preparation_workbook = load_workbook(template_preparation_output)
    template_preparation_worksheet = template_preparation_workbook["SilverTemplatePreparation"]
    template_preparation_headers = [cell.value for cell in template_preparation_worksheet[1]]

    assert template_preparation_headers == SILVER_TEMPLATE_PREPARATION_COLUMNS
    assert template_preparation_worksheet.cell(row=2, column=template_preparation_headers.index("Energy KPI") + 1).value == "Fossil Fuels"
    assert template_preparation_worksheet.cell(row=3, column=template_preparation_headers.index("Energy KPI") + 1).value == "Renewable sources"
    assert template_preparation_worksheet.cell(row=4, column=template_preparation_headers.index("Energy KPI") + 1).value == "Nuclear sources"
    assert template_preparation_worksheet.cell(row=5, column=template_preparation_headers.index("Energy KPI") + 1).value == "Renewable energy production"
    assert template_preparation_worksheet.cell(row=6, column=template_preparation_headers.index("Energy KPI") + 1).value == "Consumption of self-generated non-fuel renewable energy"
    assert template_preparation_worksheet.cell(row=2, column=template_preparation_headers.index("Total amount of energy consumed") + 1).value == pytest.approx(2592.01)
    assert template_preparation_worksheet.cell(row=2, column=template_preparation_headers.index("Amount of energy consumed") + 1).value == pytest.approx(2592.01 * 0.294)
    assert template_preparation_worksheet.cell(row=2, column=template_preparation_headers.index("contractual_instruments") + 1).value == "Contracts(e.g. PPAs)"
    assert manual_proration_calculation_output.exists()
    assert manual_proration_split_output.exists()
    assert manual_prorated_output.exists()

    manual_business_rows = [
        dict(zip(business_mapping_headers, row))
        for row in business_mapping_worksheet.iter_rows(min_row=2, values_only=True)
        if row[business_mapping_headers.index(MANUAL_DATA_ENTRY_HELPER_COLUMN)] == MANUAL_DATA_ENTRY_HELPER_VALUE
    ]
    assert {row["lookup_key"] for row in manual_business_rows} == {
        "manual unit_manual supplier_2025-06-01",
        "manual unit_manual supplier_2025-07-01",
    }
    assert {row["KPI Component"] for row in manual_business_rows} == {"Purchased electricity"}
    assert {row["Energy Type"] for row in manual_business_rows} == {"Electricity"}
    assert {row["Purchased or acquired"] for row in manual_business_rows} == {"Purchased"}
    assert {row["contractual_instruments"] for row in manual_business_rows} == {"Manual green tariff"}

    assert template_output_output.exists()
    template_output_workbook = load_workbook(template_output_output)
    template_output_worksheet = template_output_workbook["SilverTemplateOutput"]
    template_output_headers = [cell.value for cell in template_output_worksheet[1]]
    assert template_output_headers == SILVER_TEMPLATE_PREPARATION_COLUMNS
    template_output_rows = [
        dict(zip(template_output_headers, row))
        for row in template_output_worksheet.iter_rows(min_row=2, values_only=True)
    ]
    manual_template_rows = [
        row for row in template_output_rows
        if row[MANUAL_DATA_ENTRY_HELPER_COLUMN] == MANUAL_DATA_ENTRY_HELPER_VALUE
    ]
    assert len(manual_template_rows) == 6
    manual_template_amounts = {
        (_as_date(row["Consumption start date"]), row["Energy KPI"]): row["Amount of energy consumed"]
        for row in manual_template_rows
    }
    assert manual_template_amounts == {
        (date(2025, 6, 1), "Fossil Fuels"): 32,
        (date(2025, 6, 1), "Renewable sources"): 112,
        (date(2025, 6, 1), "Nuclear sources"): 16,
        (date(2025, 7, 1), "Fossil Fuels"): 35,
        (date(2025, 7, 1), "Renewable sources"): 91,
        (date(2025, 7, 1), "Nuclear sources"): 14,
    }
    template_output_amounts = [
        row[template_output_headers.index("Amount of energy consumed")]
        for row in template_output_worksheet.iter_rows(min_row=2, values_only=True)
    ]
    assert 0 not in template_output_amounts

    assert not (checkpoint_dir / "step_1_field_quality_checkpoint.csv").exists()


def test_build_silver_curated_selects_requested_columns_in_order():
    normalized_df = pd.DataFrame([
        {
            "division": "Tanker",
            "legal_entity": "Mapped Legal Entity",
            "unit_name": "Mapped Unit",
            "account_number": "ACC-001",
            "supplier_name": "Mapped Supplier",
            "consumption_start_date_1_normalized": "2025-07-01",
            "consumption_end_date_1_normalized": "2025-07-31",
            "consumption_start_date_2_normalized": "2025-08-01",
            "consumption_end_date_2_normalized": "2025-08-31",
            "quantity_1": "1234.56",
            "quantity_2": "2345.67",
            "quantity_3": "3456.78",
            "quantity_4": "4567.89",
            "quantity_5": "5678.90",
            "quantity_6": "6789.01",
            "quantity_unit_1": "KWH",
            "consumption_unit": "kWh",
            "consumption_quantity_unit": "KWH",
            "quantity_unit_2": "MWH",
            "total_amount": "7890.12",
            "solar_export": "10",
            "solar_banking_charge": "20",
            "solar_total_generation": "30",
            "division_shorthand": "TNK",
            "facility_type": "Terminal",
            "scope": "Scope 2",
            "facility_identifier": "FAC-001",
            "activity_group": "Electricity",
            "invoice_count": 1,
            "source_file": "not-curated.json",
            "sharepoint_link": "https://example.sharepoint.com/invoice.pdf",
            "status": "succeeded",
            "createdDateTime": "2026-06-14T15:07:06Z",
            "lastUpdatedDateTime": "2026-06-14T15:07:12Z",
            "apiVersion": "2024-11-30",
            "modelId": "model-one",
            "missing_fields": "quantity_4",
            "low_confidence_fields": "quantity_5",
            "approval_status": "AUTO_APPROVED",
            "manual_review_tag": "MANUALLY_REVIEWED",
        }
    ])

    curated_df = build_silver_curated(normalized_df)

    assert list(curated_df.columns) == SILVER_CURATED_COLUMNS
    assert curated_df.columns[curated_df.columns.get_loc("data_quality") + 1] == "division"
    assert curated_df.loc[0, "data_quality"] == "Actual"
    assert curated_df.loc[0, "quantity_1"] == "1234.56"
    assert curated_df.columns[curated_df.columns.get_loc("unit_name") + 1] == "account_number"
    assert curated_df.loc[0, "account_number"] == "ACC-001"
    assert curated_df.columns[curated_df.columns.get_loc("quantity_unit_1") + 1] == "consumption_unit"
    assert curated_df.columns[curated_df.columns.get_loc("consumption_unit") + 1] == "consumption_quantity_unit"
    assert curated_df.loc[0, "consumption_quantity_unit"] == "KWH"
    assert curated_df.columns[curated_df.columns.get_loc("quantity_unit_2") + 1] == "total_amount"
    assert curated_df.loc[0, "total_amount"] == "7890.12"
    assert curated_df.columns[curated_df.columns.get_loc("invoice_count") + 1] == "source_file"
    assert curated_df.columns[curated_df.columns.get_loc("source_file") + 1] == "sharepoint_link"
    assert curated_df.loc[0, "source_file"] == "not-curated.json"
    assert curated_df.loc[0, "sharepoint_link"] == "https://example.sharepoint.com/invoice.pdf"
    assert curated_df.loc[0, "createdDateTme"] == "2026-06-14T15:07:06Z"
    assert curated_df.loc[0, "lastUpdateDateTime"] == "2026-06-14T15:07:12Z"
    assert curated_df.loc[0, "modelID"] == "model-one"


def test_add_consumption_quantity_unit_prefers_quantity_unit_and_falls_back_to_mapping_unit():
    normalized_df = pd.DataFrame([
        {
            "quantity_unit_1": "KWH",
            "consumption_unit": "MWh",
        },
        {
            "quantity_unit_1": "",
            "consumption_unit": "kWh",
        },
        {
            "quantity_unit_1": "litres",
            "consumption_unit": "megawatt hours",
        },
    ])

    result = add_consumption_quantity_unit(normalized_df)

    assert result.loc[0, "consumption_quantity_unit"] == "KWH"
    assert result.loc[1, "consumption_quantity_unit"] == "KWH"
    assert result.loc[2, "consumption_quantity_unit"] == "MWH"


def test_build_silver_aggregated_sums_quantities_when_second_period_is_empty():
    curated_df = pd.DataFrame([
        {
            "division": "Tanker",
            "legal_entity": "Mapped Legal Entity",
            "unit_name": "Mapped Unit",
            "supplier_name": "Mapped Supplier",
            "consumption_start_date_1_normalized": "2025-07-01",
            "consumption_end_date_1_normalized": "2025-07-31",
            "consumption_start_date_2_normalized": "",
            "consumption_end_date_2_normalized": "",
            "quantity_1": "100.5",
            "quantity_2": 200,
            "quantity_3": "",
            "quantity_4": None,
            "quantity_5": "300.25",
            "quantity_6": "not a number",
            "quantity_unit_1": "KWH",
            "consumption_quantity_unit": "KWH",
            "quantity_unit_2": "",
        }
    ])

    aggregated_df = build_silver_aggregated(curated_df)

    assert list(aggregated_df.columns) == SILVER_AGGREGATED_COLUMNS
    assert "quantity_1" not in aggregated_df.columns
    assert aggregated_df.loc[0, "aggregated_quantity_1"] == 600.75
    assert aggregated_df.loc[0, "aggregated_quantity_2"] == ""
    assert aggregated_df.loc[0, "quantity_unit_1"] == "KWH"
    assert aggregated_df.loc[0, "consumption_quantity_unit"] == "KWH"


def test_build_silver_aggregated_keeps_first_two_quantities_when_second_period_exists():
    curated_df = pd.DataFrame([
        {
            "consumption_start_date_2_normalized": "2025-08-01",
            "consumption_end_date_2_normalized": "2025-08-31",
            "quantity_1": "1234.56",
            "quantity_2": "2345.67",
            "quantity_3": "3456.78",
            "quantity_4": "4567.89",
            "quantity_5": "5678.90",
            "quantity_6": "6789.01",
            "quantity_unit_1": "KWH",
            "consumption_quantity_unit": "KWH",
            "quantity_unit_2": "MWH",
        }
    ])

    aggregated_df = build_silver_aggregated(curated_df)

    assert aggregated_df.loc[0, "aggregated_quantity_1"] == "1234.56"
    assert aggregated_df.loc[0, "aggregated_quantity_2"] == "2345.67"
    assert aggregated_df.loc[0, "quantity_unit_2"] == "MWH"


def test_build_silver_proration_calculation_splits_quantity_across_months():
    aggregated_df = pd.DataFrame([
        {
            "source_file": "split.json",
            "division": "Tanker",
            "account_number": "ACC-001",
            "unit_name": "Unit One",
            "consumption_start_date_1_normalized": "2025-06-15",
            "consumption_end_date_1_normalized": "2025-07-20",
            "consumption_start_date_2_normalized": "",
            "consumption_end_date_2_normalized": "",
            "aggregated_quantity_1": 162700,
            "aggregated_quantity_2": "",
            "quantity_unit_1": "KWH",
            "consumption_quantity_unit": "KWH",
        }
    ])

    prorated_df = build_silver_proration_calculation(aggregated_df)

    assert list(prorated_df.columns) == SILVER_PRORATION_CALCULATION_COLUMNS
    helper_start = prorated_df.columns.get_loc("aggregated_quantity_2") + 1
    assert list(prorated_df.columns[helper_start:helper_start + 15]) == [
        "proration_source_row_id",
        "average_consumption_1",
        "date_1_proration",
        "date_1_total_days",
        "prorated_quantity_1",
        "date_1_proration_month",
        "date_1_proration_start",
        "date_1_proration_end",
        "average_consumption_date_2",
        "date_2_proration",
        "date_2_total_days",
        "prorated_quantity_2",
        "date_2_proration_month",
        "date_2_proration_start",
        "date_2_proration_end",
    ]
    assert len(prorated_df) == 2
    assert prorated_df.loc[0, "source_file"] == "split.json"
    assert prorated_df.loc[0, "division"] == "Tanker"
    assert prorated_df.loc[0, "account_number"] == "ACC-001"
    assert prorated_df.loc[0, "unit_name"] == "Unit One"
    assert prorated_df.loc[0, "consumption_start_date_1_normalized"] == "2025-06-15"
    assert prorated_df.loc[0, "consumption_end_date_1_normalized"] == "2025-07-20"
    assert prorated_df.loc[0, "aggregated_quantity_1"] == 162700
    assert prorated_df.loc[0, "average_consumption_1"] == 162700 / 36
    assert prorated_df.loc[0, "date_1_proration"] == 16
    assert prorated_df.loc[0, "date_1_total_days"] == 36
    assert prorated_df.loc[0, "prorated_quantity_1"] == pytest.approx((162700 / 36) * 16)
    assert prorated_df.loc[0, "date_1_proration_month"] == "2025-06"
    assert prorated_df.loc[0, "date_1_proration_start"] == "2025-06-15"
    assert prorated_df.loc[0, "date_1_proration_end"] == "2025-06-30"
    assert prorated_df.loc[1, "consumption_start_date_1_normalized"] == "2025-06-15"
    assert prorated_df.loc[1, "consumption_end_date_1_normalized"] == "2025-07-20"
    assert prorated_df.loc[1, "aggregated_quantity_1"] == 162700
    assert prorated_df.loc[1, "average_consumption_1"] == 162700 / 36
    assert prorated_df.loc[1, "date_1_proration"] == 20
    assert prorated_df.loc[1, "date_1_total_days"] == 36
    assert prorated_df.loc[1, "prorated_quantity_1"] == pytest.approx((162700 / 36) * 20)
    assert prorated_df.loc[1, "date_1_proration_month"] == "2025-07"
    assert prorated_df.loc[1, "date_1_proration_start"] == "2025-07-01"
    assert prorated_df.loc[1, "date_1_proration_end"] == "2025-07-20"


def test_build_silver_proration_calculation_splits_second_consumption_period():
    aggregated_df = pd.DataFrame([
        {
            "source_file": "split-second.json",
            "consumption_start_date_1_normalized": "2025-06-01",
            "consumption_end_date_1_normalized": "2025-06-30",
            "consumption_start_date_2_normalized": "2025-07-15",
            "consumption_end_date_2_normalized": "2025-08-05",
            "aggregated_quantity_1": 500,
            "aggregated_quantity_2": 210,
        }
    ])

    prorated_df = build_silver_proration_calculation(aggregated_df)

    assert len(prorated_df) == 2
    assert prorated_df.loc[0, "aggregated_quantity_1"] == 500
    assert prorated_df.loc[0, "average_consumption_1"] == 500 / 30
    assert prorated_df.loc[0, "date_1_proration"] == 30
    assert prorated_df.loc[0, "date_1_total_days"] == 30
    assert prorated_df.loc[0, "prorated_quantity_1"] == pytest.approx(500)
    assert prorated_df.loc[0, "consumption_start_date_2_normalized"] == "2025-07-15"
    assert prorated_df.loc[0, "consumption_end_date_2_normalized"] == "2025-08-05"
    assert prorated_df.loc[0, "aggregated_quantity_2"] == 210
    assert prorated_df.loc[0, "average_consumption_date_2"] == 210 / 22
    assert prorated_df.loc[0, "date_2_proration"] == 17
    assert prorated_df.loc[0, "date_2_total_days"] == 22
    assert prorated_df.loc[0, "prorated_quantity_2"] == pytest.approx((210 / 22) * 17)
    assert prorated_df.loc[1, "consumption_start_date_2_normalized"] == "2025-07-15"
    assert prorated_df.loc[1, "consumption_end_date_2_normalized"] == "2025-08-05"
    assert prorated_df.loc[1, "aggregated_quantity_2"] == 210
    assert prorated_df.loc[1, "average_consumption_date_2"] == 210 / 22
    assert prorated_df.loc[1, "date_2_proration"] == 5
    assert prorated_df.loc[1, "date_2_total_days"] == 22
    assert prorated_df.loc[1, "prorated_quantity_2"] == pytest.approx((210 / 22) * 5)


def test_build_silver_proration_calculation_counts_same_day_period_as_one_day():
    aggregated_df = pd.DataFrame([
        {
            "source_file": "same-day.json",
            "consumption_start_date_1_normalized": "2025-07-01",
            "consumption_end_date_1_normalized": "2025-07-01",
            "consumption_start_date_2_normalized": "",
            "consumption_end_date_2_normalized": "",
            "aggregated_quantity_1": 10,
            "aggregated_quantity_2": "",
        }
    ])

    prorated_df = build_silver_proration_calculation(aggregated_df)

    assert len(prorated_df) == 1
    assert prorated_df.loc[0, "average_consumption_1"] == 10
    assert prorated_df.loc[0, "date_1_proration"] == 1
    assert prorated_df.loc[0, "date_1_total_days"] == 1
    assert prorated_df.loc[0, "prorated_quantity_1"] == pytest.approx(10)


def test_build_silver_proration_split_stacks_consumption_period_columns():
    aggregated_df = pd.DataFrame([
        {
            "source_file": "split-second.json",
            "account_number": "ACC-001",
            "unit_name": "Unit One",
            "consumption_start_date_1_normalized": "2025-06-15",
            "consumption_end_date_1_normalized": "2025-07-20",
            "consumption_start_date_2_normalized": "2025-08-01",
            "consumption_end_date_2_normalized": "2025-08-31",
            "aggregated_quantity_1": 162700,
            "aggregated_quantity_2": 310,
            "quantity_unit_1": "KWH",
            "consumption_quantity_unit": "KWH",
            "quantity_unit_2": "MWH",
        }
    ])
    calculation_df = build_silver_proration_calculation(aggregated_df)

    split_df = build_silver_proration_split(calculation_df)

    assert list(split_df.columns) == SILVER_PRORATION_SPLIT_COLUMNS
    assert "consumption_start_date_1_normalized" not in split_df.columns
    assert "consumption_start_date_2_normalized" not in split_df.columns
    assert "aggregated_quantity_1" not in split_df.columns
    assert "aggregated_quantity_2" not in split_df.columns
    assert len(split_df) == 3
    assert list(split_df["date_month_proration"]) == ["2025-06", "2025-07", "2025-08"]
    assert split_df.loc[0, "consumption_start_date_proration"] == "2025-06-15"
    assert split_df.loc[0, "consumption_end_date_proration"] == "2025-07-20"
    assert split_df.loc[0, "aggregate_quantity1_to_6"] == 162700
    assert split_df.loc[0, "average_consumption"] == 162700 / 36
    assert split_df.loc[0, "date_split_days_proration"] == 16
    assert split_df.loc[0, "date_total_days_proration"] == 36
    assert split_df.loc[0, "quantity_proration"] == pytest.approx((162700 / 36) * 16)
    assert split_df.loc[0, "date_split_start_proration"] == "2025-06-15"
    assert split_df.loc[0, "date_split_end_proration"] == "2025-06-30"
    assert split_df.loc[2, "consumption_start_date_proration"] == "2025-08-01"
    assert split_df.loc[2, "consumption_end_date_proration"] == "2025-08-31"
    assert split_df.loc[2, "aggregate_quantity1_to_6"] == 310
    assert split_df.loc[2, "average_consumption"] == 310 / 31
    assert split_df.loc[2, "date_split_days_proration"] == 31
    assert split_df.loc[2, "date_total_days_proration"] == 31
    assert split_df.loc[2, "quantity_proration"] == pytest.approx(310)
    assert split_df.loc[2, "consumption_quantity_unit"] == "MWH"


def test_build_silver_prorated_aggregates_monthly_consumption_and_completeness():
    aggregated_df = pd.DataFrame([
        {
            "source_file": "june.json",
            "division": "Tanker",
            "legal_entity": "Legal One",
            "unit_name": "Unit One",
            "account_number": "ACC-001",
            "supplier_name": "Supplier One",
            "sharepoint_link": "https://example.com/june",
            "consumption_start_date_1_normalized": "2025-06-01",
            "consumption_end_date_1_normalized": "2025-06-30",
            "consumption_start_date_2_normalized": "",
            "consumption_end_date_2_normalized": "",
            "aggregated_quantity_1": 300,
            "aggregated_quantity_2": "",
            "quantity_unit_1": "KWH",
            "consumption_quantity_unit": "KWH",
        },
        {
            "source_file": "june-same-unit.json",
            "division": "Tanker",
            "legal_entity": "Legal One",
            "unit_name": "Unit One",
            "account_number": "ACC-001",
            "supplier_name": "Supplier One",
            "sharepoint_link": "https://example.com/june-same-unit",
            "consumption_start_date_1_normalized": "2025-06-01",
            "consumption_end_date_1_normalized": "2025-06-30",
            "consumption_start_date_2_normalized": "",
            "consumption_end_date_2_normalized": "",
            "aggregated_quantity_1": 150,
            "aggregated_quantity_2": "",
            "quantity_unit_1": "kWh",
            "consumption_quantity_unit": "KWH",
        },
        {
            "source_file": "july.json",
            "division": "Tanker",
            "legal_entity": "Legal One",
            "unit_name": "Unit One",
            "account_number": "ACC-001",
            "supplier_name": "Supplier One",
            "sharepoint_link": "https://example.com/july",
            "consumption_start_date_1_normalized": "2025-07-15",
            "consumption_end_date_1_normalized": "2025-07-20",
            "consumption_start_date_2_normalized": "",
            "consumption_end_date_2_normalized": "",
            "aggregated_quantity_1": 50,
            "aggregated_quantity_2": "",
            "quantity_unit_1": "KWH",
            "consumption_quantity_unit": "KWH",
        },
        {
            "source_file": "other-account.json",
            "division": "Tanker",
            "legal_entity": "Legal One",
            "unit_name": "Unit One",
            "account_number": "ACC-002",
            "supplier_name": "Supplier One",
            "sharepoint_link": "https://example.com/other-account",
            "consumption_start_date_1_normalized": "2025-06-01",
            "consumption_end_date_1_normalized": "2025-06-30",
            "consumption_start_date_2_normalized": "",
            "consumption_end_date_2_normalized": "",
            "aggregated_quantity_1": 100,
            "aggregated_quantity_2": "",
            "quantity_unit_1": "KWH",
            "consumption_quantity_unit": "KWH",
        },
    ])
    calculation_df = build_silver_proration_calculation(aggregated_df)
    split_df = build_silver_proration_split(calculation_df)

    prorated_df = build_silver_prorated(split_df)

    assert _without_numbered_sharepoint_columns(prorated_df.columns) == SILVER_PRORATED_COLUMNS
    assert "sharepoint_link_1" in prorated_df.columns
    assert "sharepoint_link_2" in prorated_df.columns
    assert (
        _without_numbered_sharepoint_columns(prorated_df.columns[:len(SILVER_PRORATION_SPLIT_COLUMNS) + 2])
        == SILVER_PRORATION_SPLIT_COLUMNS
    )
    acc_one_june = prorated_df[
        (prorated_df["account_number"] == "ACC-001")
        & (prorated_df["unit_name"] == "Unit One")
        & (prorated_df["consumption_month"] == "2025-06")
    ]
    assert len(acc_one_june) == 1
    acc_one_june = acc_one_june.iloc[0]
    assert acc_one_june["consumption_start_date_proration"] == "2025-06-01"
    assert acc_one_june["consumption_end_date_proration"] == "2025-06-30"
    assert acc_one_june["aggregate_quantity1_to_6"] == 450
    assert acc_one_june["monthly_consumption"] == pytest.approx(450)
    assert acc_one_june["consumption_quantity_unit"] == "KWH"
    assert acc_one_june["complete_month_data_captured"] == True
    assert acc_one_june["covered_days_in_month"] == 30
    assert acc_one_june["calendar_days_in_month"] == 30
    assert acc_one_june["proration_component_count"] == 2
    assert acc_one_june["proration_source_row_ids"] == "1; 2"
    assert acc_one_june["proration_split_keys"] == "1|2025-06|2025-06-01|2025-06-30; 2|2025-06|2025-06-01|2025-06-30"
    assert acc_one_june["source_file_count"] == 2
    assert acc_one_june["source_files"] == "june-same-unit.json; june.json"
    assert acc_one_june["sharepoint_link_1"] == "https://example.com/june"
    assert acc_one_june["sharepoint_link_2"] == "https://example.com/june-same-unit"

    acc_one_july = prorated_df[
        (prorated_df["account_number"] == "ACC-001")
        & (prorated_df["unit_name"] == "Unit One")
        & (prorated_df["consumption_month"] == "2025-07")
    ].iloc[0]
    assert acc_one_july["consumption_start_date_proration"] == "2025-07-15"
    assert acc_one_july["consumption_end_date_proration"] == "2025-07-20"
    assert acc_one_july["aggregate_quantity1_to_6"] == 50
    assert acc_one_july["monthly_consumption"] == pytest.approx(50)
    assert acc_one_july["complete_month_data_captured"] == False
    assert acc_one_july["covered_days_in_month"] == 6
    assert acc_one_july["calendar_days_in_month"] == 31
    assert acc_one_july["proration_component_count"] == 1
    assert acc_one_july["proration_source_row_ids"] == "3"
    assert acc_one_july["proration_split_keys"] == "3|2025-07|2025-07-15|2025-07-20"
    assert acc_one_july["sharepoint_link_1"] == "https://example.com/july"
    assert acc_one_july["sharepoint_link_2"] == ""

    assert len(prorated_df[prorated_df["account_number"] == "ACC-002"]) == 1


def test_build_silver_prorated_aggregates_second_period_by_account_number_unit_and_month():
    aggregated_df = pd.DataFrame([
        {
            "source_file": "account-one-unit-one.json",
            "division": "Tanker",
            "legal_entity": "Legal One",
            "unit_name": "Unit One",
            "account_number": "ACC-001",
            "supplier_name": "Supplier One",
            "consumption_start_date_1_normalized": "2025-07-01",
            "consumption_end_date_1_normalized": "2025-07-31",
            "consumption_start_date_2_normalized": "2025-08-01",
            "consumption_end_date_2_normalized": "2025-08-31",
            "aggregated_quantity_1": 100,
            "aggregated_quantity_2": 310,
            "quantity_unit_1": "KWH",
            "consumption_quantity_unit": "KWH",
            "quantity_unit_2": "KWH",
        },
        {
            "source_file": "account-one-unit-two.json",
            "division": "Tanker",
            "legal_entity": "Legal One",
            "unit_name": "Unit Two",
            "account_number": "ACC-001",
            "supplier_name": "Supplier One",
            "consumption_start_date_1_normalized": "2025-07-01",
            "consumption_end_date_1_normalized": "2025-07-31",
            "consumption_start_date_2_normalized": "2025-08-01",
            "consumption_end_date_2_normalized": "2025-08-31",
            "aggregated_quantity_1": 50,
            "aggregated_quantity_2": 90,
            "quantity_unit_1": "KWH",
            "consumption_quantity_unit": "KWH",
            "quantity_unit_2": "KWH",
        },
    ])
    calculation_df = build_silver_proration_calculation(aggregated_df)
    split_df = build_silver_proration_split(calculation_df)

    prorated_df = build_silver_prorated(split_df)

    august_row = prorated_df[
        (prorated_df["account_number"] == "ACC-001")
        & (prorated_df["unit_name"] == "Unit One")
        & (prorated_df["consumption_month"] == "2025-08")
    ].iloc[0]
    august_rows = prorated_df[prorated_df["consumption_month"] == "2025-08"]
    assert len(august_rows) == 2
    assert set(august_rows["unit_name"]) == {"Unit One", "Unit Two"}
    assert august_row["source_file_count"] == 1
    assert august_row["aggregate_quantity1_to_6"] == 310
    assert august_row["quantity_proration"] == pytest.approx(310)
    assert august_row["monthly_consumption"] == pytest.approx(310)
    assert august_row["consumption_quantity_unit"] == "KWH"


def test_build_silver_business_mapping_selects_template_columns_and_calculates_fields(tmp_path):
    config_dir = tmp_path / "config"
    mappings_dir = tmp_path / "mappings"
    config_dir.mkdir()
    mappings_dir.mkdir()
    (config_dir / "mapping_files.yaml").write_text(
        "\n".join([
            "mapping_files:",
            "  energy_source_allocation:",
            f"    path: {mappings_dir.joinpath('energy_source_allocation.xlsx').as_posix()}",
            "    sheet_name: energy_source_allocation",
            "    delimiter: _",
            "    ocr_key_fields: [unit_name, supplier, start_date]",
            "    workbook_key_columns: [unit_name, supplier_name, start_date]",
            "    date_key_fields: [start_date]",
            "  contracts:",
            f"    path: {mappings_dir.joinpath('contracts.xlsx').as_posix()}",
            "    sheet_name: contracts",
            "    delimiter: _",
            "    ocr_key_fields: [unit_name, supplier, contract_start_date]",
            "    workbook_key_columns: [unit_name, supplier_name, contract_start_date]",
            "    date_key_fields: [contract_start_date]",
            "",
        ]),
        encoding="utf-8",
    )
    pd.DataFrame([
        {
            "lookup_key": "unit one_supplier one_2025-06-01",
            "active": "Yes",
            "unit_name": "Unit One",
            "supplier_name": "Supplier One",
            "start_date": "2025-06-01",
            "fossil_fuel_%": 0.1,
            "renewable_energy_%": 0.8,
            "nuclear_%": 0.1,
        }
    ]).to_excel(
        mappings_dir / "energy_source_allocation.xlsx",
        sheet_name="energy_source_allocation",
        index=False,
    )
    pd.DataFrame([
        {
            "lookup_key": "unit one_supplier one_2025-06-01",
            "active": "Yes",
            "unit_name": "Unit One",
            "supplier_name": "Supplier One",
            "contract_start_date": "2025-06-01",
            "contractual_instruments": "Green tariff",
        }
    ]).to_excel(
        mappings_dir / "contracts.xlsx",
        sheet_name="contracts",
        index=False,
    )
    prorated_df = pd.DataFrame([
        {
            "data_quality": "Actual",
            "activity_group": "Electricity",
            "division": "Tanker",
            "legal_entity": "Legal One",
            "unit_name": "Unit One",
            "month_start_date": date(2025, 6, 1),
            "month_end_date": date(2025, 6, 30),
            "createdDateTme": "2026-06-14T15:07:06Z",
            "quantity_proration": 450,
            "consumption_quantity_unit": "KWH",
            "account_number": "ACC-001",
            "supplier_name": "Supplier One",
            "solar_export": "10",
            "solar_banking_charge": "5",
            "solar_total_generation": "30",
            "proration_component_count": 2,
            "proration_source_row_ids": "1; 2",
            "proration_split_keys": "1|2025-06|2025-06-01|2025-06-30; 2|2025-06|2025-06-01|2025-06-30",
            "source_file_count": 2,
            "source_files": "one.json; two.json",
            "covered_days_in_month": 30,
            "calendar_days_in_month": 30,
            "complete_month_data_captured": True,
            "status": "succeeded",
            "lastUpdateDateTime": "2026-06-14T15:07:12Z",
            "apiVersion": "2024-11-30",
            "modelID": "model-one",
            "missing_fields": "",
            "low_confidence_fields": "quantity_5",
            "approval_status": "AUTO_APPROVED",
            "manual_review_tag": "",
            "sharepoint_link_1": "https://example.com/one",
            "sharepoint_link_2": "https://example.com/two",
            "sharepoint_link_3": "https://example.com/three",
        }
    ])

    business_mapping_df = build_silver_business_mapping(prorated_df, config_dir=config_dir)

    assert list(business_mapping_df.columns) == (
        SILVER_BUSINESS_MAPPING_FIXED_COLUMNS
        + ["sharepoint_link_1", "sharepoint_link_2", "sharepoint_link_3"]
    )
    row = business_mapping_df.iloc[0]
    assert row["Data Quality"] == "Actual"
    assert row["KPI Component"] == "Purchased electricity"
    assert row["Division"] == "Tanker"
    assert row["Legal Entity Name"] == "Legal One"
    assert row["Unit"] == "Unit One"
    assert row["Consumption start date"] == "2025-06-01"
    assert row["Consumption end date"] == "2025-06-30"
    assert row["Transaction Date"] == "2026-06-14"
    assert row["Energy Type"] == "Electricity"
    assert row["Purchased or acquired"] == "Purchased"
    assert row["Amount of energy consumed"] == 450
    assert row["Energy Unit"] == "KWH"
    assert row["fossil_fuel_%"] == 0.1
    assert row["renewable_energy_%"] == 0.8
    assert row["nuclear_%"] == 0.1
    assert row["contractual_instruments"] == "Green tariff"
    assert row["Account number"] == "ACC-001"
    assert row["Supplier Name"] == "Supplier One"
    assert row["lookup_key"] == "unit one_supplier one_2025-06-01"
    assert row["total_solar_consumed"] == 20
    assert row["sharepoint_link_1"] == "https://example.com/one"
    assert row["sharepoint_link_2"] == "https://example.com/two"
    assert row["sharepoint_link_3"] == "https://example.com/three"


def test_build_silver_template_preparation_unpivots_energy_kpis_from_business_mapping():
    business_mapping_df = pd.DataFrame([
        {
            "Data Quality": "Actual",
            "KPI Component": "Purchased electricity",
            "Division": "Tanker",
            "Legal Entity Name": "Legal One",
            "Unit": "Unit One",
            "Consumption start date": "2025-06-01",
            "Consumption end date": "2025-06-30",
            "Transaction Date": "2026-06-14",
            "Energy Type": "Electricity",
            "Purchased or acquired": "Purchased",
            "Amount of energy consumed": 400,
            "Energy Unit": "KWH",
            "lookup_key": "unit one_supplier one_2025-06-01",
            "fossil_fuel_%": 0.25,
            "renewable_energy_%": 0.5,
            "nuclear_%": 0.25,
            "contractual_instruments": "Green tariff",
            "Account number": "ACC-001",
            "Supplier Name": "Supplier One",
            "solar_export": "10",
            "solar_banking_charge": "5",
            "solar_total_generation": "30",
            "total_solar_consumed": 20,
            "proration_component_count": 2,
            "proration_source_row_ids": "1; 2",
            "proration_split_keys": "1|2025-06|2025-06-01|2025-06-30",
            "source_file_count": 2,
            "source_files": "one.json; two.json",
            "covered_days_in_month": 30,
            "calendar_days_in_month": 30,
            "complete_month_data_captured": True,
            "status": "succeeded",
            "lastUpdateDateTime": "2026-06-14T15:07:12Z",
            "apiVersion": "2024-11-30",
            "modelID": "model-one",
            "missing_fields": "",
            "low_confidence_fields": "",
            "approval_status": "AUTO_APPROVED",
            "manual_review_tag": "",
            "sharepoint_link_1": "https://example.com/one",
            "sharepoint_link_2": "https://example.com/two",
        }
    ])

    template_df = build_silver_template_preparation(business_mapping_df)

    assert list(template_df.columns) == SILVER_TEMPLATE_PREPARATION_COLUMNS
    assert list(template_df["Energy KPI"]) == [
        "Fossil Fuels",
        "Renewable sources",
        "Nuclear sources",
        "Renewable energy production",
        "Consumption of self-generated non-fuel renewable energy",
    ]
    assert list(template_df["Amount of energy consumed"]) == [100, 200, 100, 30, 20]
    assert list(template_df["Energy Unit"]) == ["KWH", "KWH", "KWH", "KWH", "KWH"]
    assert list(template_df["Total amount of energy consumed"]) == [400, 400, 400, 400, 400]
    assert list(template_df["fossil_fuel_%"]) == [0.25, 0.25, 0.25, 0.25, 0.25]
    assert list(template_df["renewable_energy_%"]) == [0.5, 0.5, 0.5, 0.5, 0.5]
    assert list(template_df["nuclear_%"]) == [0.25, 0.25, 0.25, 0.25, 0.25]
    assert list(template_df["solar_total_generation"]) == ["30", "30", "30", "30", "30"]
    assert list(template_df["total_solar_consumed"]) == [20, 20, 20, 20, 20]
    assert set(template_df["contractual_instruments"]) == {"Green tariff"}
    assert set(template_df["lookup_key"]) == {"unit one_supplier one_2025-06-01"}
    assert set(template_df["Transaction Date"]) == {"2026-06-14"}
    assert set(template_df["KPI Component"]) == {"Purchased electricity"}


def test_build_silver_template_output_filters_zero_amount_rows():
    template_preparation_df = pd.DataFrame([
        {
            "Energy KPI": "Fossil Fuels",
            "Amount of energy consumed": 100,
            "Energy Unit": "KWH",
        },
        {
            "Energy KPI": "Renewable sources",
            "Amount of energy consumed": 0,
            "Energy Unit": "KWH",
        },
        {
            "Energy KPI": "Nuclear sources",
            "Amount of energy consumed": "0.0",
            "Energy Unit": "KWH",
        },
        {
            "Energy KPI": "Renewable energy production",
            "Amount of energy consumed": "",
            "Energy Unit": "KWH",
        },
        {
            "Energy KPI": "Consumption of self-generated non-fuel renewable energy",
            "Amount of energy consumed": None,
            "Energy Unit": "KWH",
        },
    ]).reindex(columns=SILVER_TEMPLATE_PREPARATION_COLUMNS, fill_value="")

    template_output_df = build_silver_template_output(template_preparation_df)

    assert list(template_output_df["Energy KPI"]) == ["Fossil Fuels"]
    assert list(template_output_df["Amount of energy consumed"]) == [100]
    assert list(template_output_df.columns) == SILVER_TEMPLATE_PREPARATION_COLUMNS


def test_manual_data_entry_flow_prorates_completed_business_mapped_rows():
    manual_decisions_df = pd.DataFrame([
        {
            "invoice_id": "manual-one",
            "line_id": "1",
            "source_file": "manual.pdf",
            "manual_entry_status": "COMPLETED",
            "reviewed_by": "Reviewer",
            "reviewed_at": "2026-06-30T00:00:00+00:00",
            "data_quality": "Estimated",
            "division": "Tanker",
            "legal_entity_name": "Legal One",
            "unit": "Unit One",
            "consumption_start_date": "2025-06-15",
            "consumption_end_date": "2025-07-14",
            "transaction_date": "2025-07-20",
            "amount_of_energy_consumed": "300",
            "energy_unit": "kWh",
        }
    ])

    proration_input_df = build_manual_data_entry_proration_input(manual_decisions_df)
    calculation_df = build_silver_proration_calculation(proration_input_df)
    split_df = build_silver_proration_split(calculation_df)
    prorated_df = build_silver_prorated(split_df)

    assert list(prorated_df["month_start_date"]) == [date(2025, 6, 1), date(2025, 7, 1)]
    assert list(prorated_df["month_end_date"]) == [date(2025, 6, 30), date(2025, 7, 31)]
    assert list(prorated_df["quantity_proration"]) == [160, 140]
    assert set(prorated_df["consumption_quantity_unit"]) == {"KWH"}
    assert set(prorated_df[MANUAL_DATA_ENTRY_HELPER_COLUMN]) == {MANUAL_DATA_ENTRY_HELPER_VALUE}

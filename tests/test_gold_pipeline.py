from datetime import date

import pandas as pd
from openpyxl import load_workbook

from src.output.xlsx_writer import write_xlsx
from src.pipeline.run_gold_pipeline import build_gold_template, run_gold_pipeline
from src.pipeline.run_silver_pipeline import silver_output_filename


def _as_date(value):
    return value.date() if hasattr(value, "date") else value


def test_build_gold_template_applies_unit_conversions():
    template_output_df = pd.DataFrame([
        {
            "Energy KPI": "Fossil Fuels",
            "Amount of energy consumed": 1.25,
            "Energy Unit": "MWh",
            "Total amount of energy consumed": 2,
            "source_files": "one.json",
        },
        {
            "Energy KPI": "Renewable sources",
            "Amount of energy consumed": 50,
            "Energy Unit": "kWh",
            "Total amount of energy consumed": 100,
            "source_files": "two.json",
        },
    ])

    result = build_gold_template(template_output_df)

    assert list(result.columns) == [
        "Energy KPI",
        "Amount of energy consumed",
        "Energy Unit",
        "Total amount of energy consumed",
        "Original Energy Unit",
        "source_files",
    ]
    assert list(result["Amount of energy consumed"]) == [1250, 50]
    assert list(result["Energy Unit"]) == ["kWh", "kWh"]
    assert list(result["Original Energy Unit"]) == ["MWh", "kWh"]
    assert list(result["source_files"]) == ["one.json", "two.json"]


def test_run_gold_pipeline_formats_template_dates_like_silver_step_10(tmp_path):
    config_dir = tmp_path / "config"
    silver_dir = tmp_path / "silver"
    gold_dir = tmp_path / "gold"
    config_dir.mkdir()
    silver_dir.mkdir()
    (config_dir / "settings.yaml").write_text(
        "\n".join([
            "paths:",
            f"  silver_excel_output: {silver_dir.as_posix()}",
            f"  gold_output: {gold_dir.as_posix()}",
            "",
        ]),
        encoding="utf-8",
    )
    silver_template_output = pd.DataFrame([
        {
            "Consumption start date": "2025-06-01",
            "Consumption end date": "2025-06-30",
            "Transaction Date": "2026-06-14",
            "Energy KPI": "Fossil Fuels",
            "Amount of energy consumed": 1.25,
            "Energy Unit": "MWh",
            "Total amount of energy consumed": 2,
        }
    ])
    write_xlsx(
        silver_template_output,
        silver_dir,
        silver_output_filename("scope2", "template_output"),
        sheet_name="SilverTemplateOutput",
    )

    output = run_gold_pipeline(config_dir=config_dir)

    workbook = load_workbook(output)
    worksheet = workbook["GoldTemplate"]
    headers = [cell.value for cell in worksheet[1]]

    for header, expected_date in {
        "Consumption start date": date(2025, 6, 1),
        "Consumption end date": date(2025, 6, 30),
        "Transaction Date": date(2026, 6, 14),
    }.items():
        cell = worksheet.cell(row=2, column=headers.index(header) + 1)
        assert _as_date(cell.value) == expected_date
        assert cell.number_format == "yyyy-mm-dd"

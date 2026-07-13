from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ISO_DATE_COLUMNS = {
    "invoice_date_normalized",
    "consumption_start_date_1_normalized",
    "consumption_end_date_1_normalized",
    "consumption_start_date_2_normalized",
    "consumption_end_date_2_normalized",
    "date_1_proration_start",
    "date_1_proration_end",
    "date_2_proration_start",
    "date_2_proration_end",
    "consumption_start_date_proration",
    "consumption_end_date_proration",
    "date_split_start_proration",
    "date_split_end_proration",
    "month_start_date",
    "month_end_date",
    "Consumption start date",
    "Consumption end date",
    "Transaction Date",
}


RAW_DATE_COLUMNS = {
    "invoice_date",
    "consumption_start_date_1",
    "consumption_end_date_1",
    "consumption_start_date_2",
    "consumption_end_date_2",
}


HYPERLINK_COLUMNS = {
    "sharepoint_link",
}


NUMBER_COLUMNS = {
    "quantity_1",
    "quantity_2",
    "quantity_3",
    "quantity_4",
    "quantity_5",
    "quantity_6",
    "aggregated_quantity_1",
    "aggregated_quantity_2",
    "average_consumption_1",
    "date_1_proration",
    "date_1_total_days",
    "prorated_quantity_1",
    "average_consumption_date_2",
    "date_2_proration",
    "date_2_total_days",
    "prorated_quantity_2",
    "aggregate_quantity1_to_6",
    "average_consumption",
    "date_split_days_proration",
    "date_total_days_proration",
    "quantity_proration",
    "monthly_consumption",
    "Amount of energy consumed",
    "Total amount of energy consumed",
    "solar_export",
    "solar_banking_charge",
    "solar_total_generation",
    "total_solar_consumed",
    "fossil_fuel_%",
    "renewable_energy_%",
    "nuclear_%",
    "covered_days_in_month",
    "calendar_days_in_month",
    "proration_component_count",
    "source_file_count",
    "total_amount",
}


def write_xlsx(
    df: pd.DataFrame,
    output_dir: str | Path,
    filename: str,
    sheet_name: str = "Silver",
) -> Path:
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    target = output_path / filename

    try:
        with pd.ExcelWriter(target, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            worksheet = writer.sheets[sheet_name]
            _format_worksheet(worksheet)
    except PermissionError as exc:
        raise PermissionError(
            f"Cannot write Excel output to {target}. "
            "Close the workbook if it is open in Excel or another app, then run the pipeline again."
        ) from exc

    return target


def dataframe_to_xlsx_bytes(df: pd.DataFrame, sheet_name: str = "Silver") -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.sheets[sheet_name]
        _format_worksheet(worksheet)
    return output.getvalue()


def _format_worksheet(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    headers = [cell.value for cell in worksheet[1]]

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_index, header in enumerate(headers, start=1):
        letter = get_column_letter(column_index)
        if header in RAW_DATE_COLUMNS:
            _format_text_column(worksheet, letter)
        elif header in ISO_DATE_COLUMNS:
            _format_iso_date_column(worksheet, letter)
        elif header in HYPERLINK_COLUMNS or _is_numbered_sharepoint_link(header):
            _format_hyperlink_column(worksheet, letter)
        elif header in NUMBER_COLUMNS:
            _format_number_column(worksheet, letter)
        worksheet.column_dimensions[letter].width = _column_width(worksheet, column_index)


def _format_text_column(worksheet, column_letter: str) -> None:
    for cell in worksheet[column_letter][1:]:
        cell.number_format = "@"


def _format_iso_date_column(worksheet, column_letter: str) -> None:
    for cell in worksheet[column_letter][1:]:
        if isinstance(cell.value, str) and _is_iso_date(cell.value):
            cell.value = datetime.strptime(cell.value, "%Y-%m-%d").date()
            cell.number_format = "yyyy-mm-dd"
        elif isinstance(cell.value, datetime):
            cell.value = cell.value.date()
            cell.number_format = "yyyy-mm-dd"
        elif isinstance(cell.value, date):
            cell.number_format = "yyyy-mm-dd"
        else:
            cell.number_format = "@"


def _format_hyperlink_column(worksheet, column_letter: str) -> None:
    hyperlink_font = Font(color="0563C1", underline="single")
    for cell in worksheet[column_letter][1:]:
        if isinstance(cell.value, str) and cell.value.strip():
            cell.hyperlink = cell.value.strip()
            cell.font = hyperlink_font


def _format_number_column(worksheet, column_letter: str) -> None:
    for cell in worksheet[column_letter][1:]:
        numeric_value = _coerce_number(cell.value)
        if numeric_value is not None:
            cell.value = numeric_value
        cell.number_format = "General"


def _coerce_number(value) -> float | int | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, int | float):
        return value
    if not isinstance(value, str):
        return None

    cleaned = value.strip().replace(",", "")
    if not cleaned:
        return None
    try:
        number = float(cleaned)
    except ValueError:
        return None
    if number.is_integer():
        return int(number)
    return number


def _is_iso_date(value: str) -> bool:
    try:
        datetime.strptime(value, "%Y-%m-%d")
    except ValueError:
        return False
    return True


def _is_numbered_sharepoint_link(header) -> bool:
    if not isinstance(header, str) or not header.startswith("sharepoint_link_"):
        return False
    suffix = header.rsplit("_", 1)[1]
    return suffix.isdigit()


def _column_width(worksheet, column_index: int) -> int:
    values = [
        "" if cell.value is None else str(cell.value)
        for cell in worksheet.iter_cols(min_col=column_index, max_col=column_index, values_only=False).__next__()
    ]
    longest = max((len(value) for value in values), default=12)
    return min(max(longest + 2, 12), 42)
